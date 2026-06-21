#!/usr/bin/env python3
"""
extraer_predicciones.py
========================
Script de uso único para generar predicciones.json desde los archivos Excel.
Ejecutar: python3 extraer_predicciones.py
Requiere: pip install openpyxl

NOTA: Solo necesitas correr esto si agregas nuevos participantes.
      Para actualizar resultados, edita resultados.json directamente.
"""
import openpyxl, os, json, sys
from datetime import date

EXCEL_DIR   = "."   # Directorio con los archivos *_corregido_equipos_corregidos.xlsx
OUTPUT_FILE = "predicciones.json"

MATCH_INFO = {
    1:{"local":"México","visitante":"Sudáfrica","grupo":"A","jornada":1},
    2:{"local":"Corea del Sur","visitante":"Chequia","grupo":"A","jornada":1},
    25:{"local":"Chequia","visitante":"Sudáfrica","grupo":"A","jornada":2},
    28:{"local":"México","visitante":"Corea del Sur","grupo":"A","jornada":2},
    53:{"local":"Chequia","visitante":"México","grupo":"A","jornada":3},
    54:{"local":"Sudáfrica","visitante":"Corea del Sur","grupo":"A","jornada":3},
    3:{"local":"Canadá","visitante":"Bosnia","grupo":"B","jornada":1},
    8:{"local":"Qatar","visitante":"Suiza","grupo":"B","jornada":1},
    26:{"local":"Suiza","visitante":"Bosnia","grupo":"B","jornada":2},
    27:{"local":"Canadá","visitante":"Qatar","grupo":"B","jornada":2},
    51:{"local":"Bosnia","visitante":"Suiza","grupo":"B","jornada":3},
    52:{"local":"Qatar","visitante":"Bosnia","grupo":"B","jornada":3},
    7:{"local":"Brasil","visitante":"Marruecos","grupo":"C","jornada":1},
    5:{"local":"Haití","visitante":"Escocia","grupo":"C","jornada":1},
    30:{"local":"Escocia","visitante":"Marruecos","grupo":"C","jornada":2},
    29:{"local":"Brasil","visitante":"Haití","grupo":"C","jornada":2},
    49:{"local":"Escocia","visitante":"Brasil","grupo":"C","jornada":3},
    50:{"local":"Marruecos","visitante":"Haití","grupo":"C","jornada":3},
    4:{"local":"EE.UU.","visitante":"Paraguay","grupo":"D","jornada":1},
    6:{"local":"Australia","visitante":"Turquía","grupo":"D","jornada":1},
    32:{"local":"EE.UU.","visitante":"Australia","grupo":"D","jornada":2},
    31:{"local":"Turquía","visitante":"Paraguay","grupo":"D","jornada":2},
    59:{"local":"Turquía","visitante":"EE.UU.","grupo":"D","jornada":3},
    60:{"local":"Paraguay","visitante":"Australia","grupo":"D","jornada":3},
    10:{"local":"Alemania","visitante":"Curazao","grupo":"E","jornada":1},
    9:{"local":"C. de Marfil","visitante":"Ecuador","grupo":"E","jornada":1},
    33:{"local":"Alemania","visitante":"C. de Marfil","grupo":"E","jornada":2},
    34:{"local":"Ecuador","visitante":"Curazao","grupo":"E","jornada":2},
    55:{"local":"Curazao","visitante":"C. de Marfil","grupo":"E","jornada":3},
    56:{"local":"Ecuador","visitante":"Alemania","grupo":"E","jornada":3},
    11:{"local":"P. Bajos","visitante":"Japón","grupo":"F","jornada":1},
    12:{"local":"Suecia","visitante":"Túnez","grupo":"F","jornada":1},
    35:{"local":"P. Bajos","visitante":"Suecia","grupo":"F","jornada":2},
    36:{"local":"Túnez","visitante":"Japón","grupo":"F","jornada":2},
    57:{"local":"Japón","visitante":"Suecia","grupo":"F","jornada":3},
    58:{"local":"Túnez","visitante":"P. Bajos","grupo":"F","jornada":3},
    16:{"local":"Bélgica","visitante":"Egipto","grupo":"G","jornada":1},
    15:{"local":"Irán","visitante":"N. Zelanda","grupo":"G","jornada":1},
    39:{"local":"Bélgica","visitante":"Irán","grupo":"G","jornada":2},
    40:{"local":"N. Zelanda","visitante":"Egipto","grupo":"G","jornada":2},
    63:{"local":"Egipto","visitante":"Irán","grupo":"G","jornada":3},
    64:{"local":"N. Zelanda","visitante":"Bélgica","grupo":"G","jornada":3},
    14:{"local":"España","visitante":"Cabo Verde","grupo":"H","jornada":1},
    13:{"local":"Arabia Saudita","visitante":"Uruguay","grupo":"H","jornada":1},
    38:{"local":"España","visitante":"Arabia Saudita","grupo":"H","jornada":2},
    37:{"local":"Uruguay","visitante":"Cabo Verde","grupo":"H","jornada":2},
    65:{"local":"Cabo Verde","visitante":"Arabia Saudita","grupo":"H","jornada":3},
    66:{"local":"Uruguay","visitante":"España","grupo":"H","jornada":3},
    17:{"local":"Francia","visitante":"Senegal","grupo":"I","jornada":1},
    18:{"local":"Irak","visitante":"Noruega","grupo":"I","jornada":1},
    42:{"local":"Francia","visitante":"Irak","grupo":"I","jornada":2},
    41:{"local":"Noruega","visitante":"Senegal","grupo":"I","jornada":2},
    61:{"local":"Noruega","visitante":"Francia","grupo":"I","jornada":3},
    62:{"local":"Senegal","visitante":"Irak","grupo":"I","jornada":3},
    19:{"local":"Argentina","visitante":"Argelia","grupo":"J","jornada":1},
    20:{"local":"Austria","visitante":"Jordania","grupo":"J","jornada":1},
    43:{"local":"Argentina","visitante":"Austria","grupo":"J","jornada":2},
    44:{"local":"Jordania","visitante":"Argelia","grupo":"J","jornada":2},
    69:{"local":"Argelia","visitante":"Austria","grupo":"J","jornada":3},
    70:{"local":"Jordania","visitante":"Argentina","grupo":"J","jornada":3},
    23:{"local":"Portugal","visitante":"Congo","grupo":"K","jornada":1},
    24:{"local":"Uzbekistán","visitante":"Colombia","grupo":"K","jornada":1},
    47:{"local":"Portugal","visitante":"Uzbekistán","grupo":"K","jornada":2},
    48:{"local":"Colombia","visitante":"Congo","grupo":"K","jornada":2},
    71:{"local":"Colombia","visitante":"Portugal","grupo":"K","jornada":3},
    72:{"local":"Congo","visitante":"Uzbekistán","grupo":"K","jornada":3},
    22:{"local":"Inglaterra","visitante":"Croacia","grupo":"L","jornada":1},
    21:{"local":"Ghana","visitante":"Panamá","grupo":"L","jornada":1},
    45:{"local":"Inglaterra","visitante":"Ghana","grupo":"L","jornada":2},
    46:{"local":"Panamá","visitante":"Croacia","grupo":"L","jornada":2},
    67:{"local":"Panamá","visitante":"Inglaterra","grupo":"L","jornada":3},
    68:{"local":"Croacia","visitante":"Ghana","grupo":"L","jornada":3},
}

HONOR_ROWS_TEAM   = {150:"campeon",151:"subcampeon",152:"tercer_puesto"}
HONOR_ROWS_PLAYER = {154:"botin_oro",155:"botin_plata",156:"botin_bronce",
                     158:"balon_oro",159:"balon_plata",160:"balon_bronce"}

def extraer():
    files = sorted([f for f in os.listdir(EXCEL_DIR) if f.endswith('_corregido_equipos_corregidos.xlsx')])
    if not files:
        print(f"ERROR: No se encontraron archivos Excel en {EXCEL_DIR}")
        sys.exit(1)

    participantes = []
    errors = []

    for i, fname in enumerate(files):
        try:
            wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, fname), data_only=True)
            ws = wb['WORLDCUP']
            name = fname.replace('_corregido_equipos_corregidos.xlsx','').replace('_',' ')
            pid  = fname.replace('_corregido_equipos_corregidos.xlsx','').lower()

            predicciones = {}
            for r in range(4, 148):
                ah = ws.cell(r, 34).value
                ac = ws.cell(r, 29).value
                ad = ws.cell(r, 30).value
                if ah is None or ac is None or ad is None:
                    continue
                predicciones[f"p{int(ah):03d}"] = {
                    "golesLocal": int(ac),
                    "golesVisitante": int(ad)
                }

            honor = {}
            for row, key in HONOR_ROWS_TEAM.items():
                val = ws.cell(row, 27).value
                if val: honor[key] = str(val).strip()
            for row, key in HONOR_ROWS_PLAYER.items():
                val = ws.cell(row, 29).value
                if val: honor[key] = str(val).strip()

            participantes.append({"id":pid,"nombre":name,"predicciones":predicciones,"honor":honor})
            if (i+1) % 20 == 0:
                print(f"  Procesados: {i+1}/{len(files)}")
        except Exception as e:
            errors.append({"archivo": fname, "error": str(e)})
            print(f"  ERROR en {fname}: {e}")

    output = {
        "_meta": {
            "generado": str(date.today()),
            "total_participantes": len(participantes),
            "nota": "Generado automáticamente. No editar manualmente."
        },
        "participantes": participantes
    }

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"\n✅ {OUTPUT_FILE} generado: {len(participantes)} participantes, {size_kb:.0f} KB")
    if errors:
        print(f"⚠️  {len(errors)} errores:")
        for e in errors: print(f"   {e['archivo']}: {e['error']}")

if __name__ == "__main__":
    extraer()
